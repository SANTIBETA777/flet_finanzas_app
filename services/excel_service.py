import openpyxl
from openpyxl import Workbook

class ExcelService:
    def __init__(self):
        self.file_path = "data/gastos.xlsx"
        self._ensure_excel_exists()

    def _ensure_excel_exists(self):
        try:
            openpyxl.load_workbook(self.file_path)
        except:
            wb = Workbook()
            ws = wb.active
            ws.title = "Gastos"
            ws.append(["Fecha", "Descripción", "Monto"])
            wb.save(self.file_path)

    def add_expense(self, fecha, descripcion, monto):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active
        ws.append([fecha, descripcion, float(monto)])
        wb.save(self.file_path)

    def get_expenses(self):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        return rows[1:]
