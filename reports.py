# reports.py
import os
from datetime import datetime
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from models import obtener_transacciones


# ============================================================
#   EXPORTAR HISTORIAL COMPLETO A EXCEL
# ============================================================

def exportar_historial_excel(ruta="reportes/historial.xlsx"):
    trans = obtener_transacciones()

    os.makedirs("reportes", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"

    ws.append(["Fecha", "Tipo", "Monto", "Categoría", "Descripción"])

    for t in trans:
        ws.append([
            t.fecha,
            t.tipo,
            t.monto,
            t.categoria_nombre or "—",
            t.descripcion,
        ])

    wb.save(ruta)
    return ruta


# ============================================================
#   EXPORTAR INGRESOS/GASTOS POR RANGO DE FECHAS
# ============================================================

def exportar_por_rango_excel(fecha_desde, fecha_hasta, ruta="reportes/rango.xlsx"):
    trans = obtener_transacciones()

    filtradas = [
        t for t in trans
        if fecha_desde <= t.fecha <= fecha_hasta
    ]

    os.makedirs("reportes", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rango"

    ws.append(["Fecha", "Tipo", "Monto", "Categoría", "Descripción"])

    for t in filtradas:
        ws.append([
            t.fecha,
            t.tipo,
            t.monto,
            t.categoria_nombre or "—",
            t.descripcion,
        ])

    wb.save(ruta)
    return ruta


# ============================================================
#   EXPORTAR ESTADO DE CUENTA A EXCEL
# ============================================================

def exportar_estado_cuenta_excel(ruta="reportes/estado_cuenta.xlsx"):
    trans = obtener_transacciones()

    ingresos = sum(t.monto for t in trans if t.tipo == "ingreso")
    gastos = sum(t.monto for t in trans if t.tipo == "gasto")
    saldo = ingresos - gastos

    os.makedirs("reportes", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Cuenta"

    ws.append(["Concepto", "Valor"])
    ws.append(["Total Ingresos", ingresos])
    ws.append(["Total Gastos", gastos])
    ws.append(["Saldo Actual", saldo])

    wb.save(ruta)
    return ruta


# ============================================================
#   EXPORTAR HISTORIAL A PDF (PROFESIONAL)
# ============================================================

def exportar_historial_pdf(ruta="reportes/historial.pdf"):
    trans = obtener_transacciones()

    os.makedirs("reportes", exist_ok=True)

    c = canvas.Canvas(ruta, pagesize=letter)
    width, height = letter

    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, height - 50, "Historial de Transacciones")

    c.setFont("Helvetica", 10)
    y = height - 90

    c.drawString(50, y, "Fecha")
    c.drawString(120, y, "Tipo")
    c.drawString(180, y, "Monto")
    c.drawString(250, y, "Categoría")
    c.drawString(350, y, "Descripción")

    y -= 20

    for t in trans:
        if y < 50:  # Nueva página
            c.showPage()
            c.setFont("Helvetica", 10)
            y = height - 50

        c.drawString(50, y, t.fecha)
        c.drawString(120, y, t.tipo)
        c.drawString(180, y, f"${t.monto:.0f}")
        c.drawString(250, y, t.categoria_nombre or "—")
        c.drawString(350, y, t.descripcion[:40])

        y -= 20

    c.save()
    return ruta


# ============================================================
#   EXPORTAR ESTADO DE CUENTA A PDF
# ============================================================

def exportar_estado_cuenta_pdf(ruta="reportes/estado_cuenta.pdf"):
    trans = obtener_transacciones()

    ingresos = sum(t.monto for t in trans if t.tipo == "ingreso")
    gastos = sum(t.monto for t in trans if t.tipo == "gasto")
    saldo = ingresos - gastos

    os.makedirs("reportes", exist_ok=True)

    c = canvas.Canvas(ruta, pagesize=letter)
    width, height = letter

    c.setFont("Helvetica-Bold", 18)
    c.drawString(50, height - 50, "Estado de Cuenta")

    c.setFont("Helvetica", 12)
    y = height - 100

    c.drawString(50, y, f"Total Ingresos: ${ingresos:,.0f}")
    y -= 20
    c.drawString(50, y, f"Total Gastos: ${gastos:,.0f}")
    y -= 20
    c.drawString(50, y, f"Saldo Actual: ${saldo:,.0f}")

    c.save()
    return ruta


# ============================================================
#   FUNCIONES QUE FALTABAN PARA INGRESOSSCREEN
# ============================================================

def exportar_transacciones_excel(ruta):
    """Exporta TODAS las transacciones a Excel (compatibilidad con IngresosScreen)."""
    return exportar_historial_excel(ruta)


def exportar_transacciones_pdf(ruta):
    """Exporta TODAS las transacciones a PDF (compatibilidad con IngresosScreen)."""
    return exportar_historial_pdf(ruta)
